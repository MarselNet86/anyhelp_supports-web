from django.contrib.auth import login, logout
from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect, get_object_or_404
from openpyxl import load_workbook
from django.contrib import messages
from datetime import datetime
from decimal import Decimal, InvalidOperation
from .forms import SupportExcelUploadForm, EmailAuthForm
from .models import Support, FieldSupport
from django.views.decorators.http import require_POST
from django.http import JsonResponse, HttpResponseBadRequest, HttpResponse
from django.views.decorators.csrf import csrf_exempt
from django.core.exceptions import ValidationError
from decimal import Decimal, InvalidOperation
from django.template.loader import render_to_string
from django.core.files.storage import default_storage

from io import BytesIO
from zipfile import ZipFile, ZIP_DEFLATED
import os

from .models import FieldSupport


def login_view(request):
    if request.user.is_authenticated:
        return redirect("main:supports_upload")
    
    if request.method == "POST":
        form = EmailAuthForm(request=request, data=request.POST)
        if form.is_valid():
            login(request, form.get_user())
            return redirect("main:supports_upload")
    else:
        form = EmailAuthForm()
        
    return render(request, "main/login.html", {"form": form})

def logout_view(request):
    logout(request)
    return redirect("main:login")


@login_required
def supports_upload_view(request):
    if request.method == "GET":
        form = SupportExcelUploadForm()
        return render(request, "main/supports_upload.html", {"form": form})

    elif request.method == "POST":
        try:
            form = SupportExcelUploadForm(request.POST, request.FILES)
            if not form.is_valid():
                html = render_to_string(
                    "main/alert.html",
                    {"type": "error", "text": "Форма содержит ошибки"},
                    request=request,
                )
                return JsonResponse({
                    "success": False,
                    "error": "Форма содержит ошибки",
                    "form_errors": form.errors,
                    "alert_html": html,
                })

            file = request.FILES["file"]
            if file.size > 50 * 1024 * 1024:  # 50MB
                html = render_to_string(
                    "main/alert.html",
                    {"type": "error", "text": "Файл слишком большой. Максимальный размер: 50MB"},
                    request=request,
                )
                return JsonResponse({
                    "success": False,
                    "error": "Файл слишком большой. Максимальный размер: 50MB",
                    "alert_html": html
                })

            result = process_excel_file(file)

            if result["success"]:
                text = f"Импорт завершён. Создано: {result['created']}, обновлено: {result['updated']}."
                messages.success(request, text)  # пригодится в non-AJAX/redirect сценарии

                alert_html = render_to_string(
                    "main/alert.html",
                    {"type": "success", "text": text},
                    request=request,
                )
                return JsonResponse({
                    "success": True,
                    "created": result["created"],
                    "updated": result["updated"],
                    "total": result["created"] + result["updated"],
                    "message": text,
                    "alert_html": alert_html,
                    "errors": result.get("errors", []),
                })
            else:
                alert_html = render_to_string(
                    "main/alert.html",
                    {"type": "error", "text": result["error"]},
                    request=request,
                )
                return JsonResponse({
                    "success": False,
                    "error": result["error"],
                    "alert_html": alert_html
                })

        except Exception as e:
            err = f"Произошла ошибка при обработке файла: {str(e)}"
            alert_html = render_to_string(
                "main/alert.html", {"type": "error", "text": err}, request=request
            )
            return JsonResponse({"success": False, "error": err, "alert_html": alert_html})


def process_excel_file(file):
    """
    Обрабатывает Excel файл и возвращает результат
    """
    try:
        wb = load_workbook(file, data_only=True)
        ws = wb.active

        headers = [str(c.value).strip() if c.value else "" for c in ws[1]]

        # карта заголовков → поля модели
        header_map = {
            "Населенный пункт": "settlement",
            "Филиал": "branch",
            "Номер опоры": "support_number",
            "Название": "name",
            "Уточнение расположения(ГИД)": "address",
            "Долгота": "longitude",
            "Широта": "latitude",
            "Дата ввода в экспуатацию": "commissioning_date",
            "Владеющая организация": "owner",
            "Материал несущей конструкции": "material",
        }

        # Проверяем наличие необходимых заголовков
        missing_headers = [h for h in header_map.keys() if h not in headers]
        if missing_headers:
            return {'success': False,
                    'error': f'В файле отсутствуют обязательные колонки: {", ".join(missing_headers)}'}

        # индексы колонок
        idx = {}
        for i, h in enumerate(headers):
            if h in header_map:
                idx[header_map[h]] = i

        created = 0
        updated = 0
        errors = []

        # (информационно) — общее количество строк
        total_rows = ws.max_row - 1  # исключаем заголовок
        processed_rows = 0

        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not any(row):
                processed_rows += 1
                continue

            try:
                data = {}
                for field, i in idx.items():
                    val = row[i] if i < len(row) else None

                    if field in ("longitude", "latitude"):
                        try:
                            data[field] = Decimal(str(val).replace(",", ".")) if val else None
                        except (InvalidOperation, AttributeError):
                            data[field] = None
                    elif field == "commissioning_date":
                        if isinstance(val, datetime):
                            data[field] = val
                        elif isinstance(val, str) and val.strip():
                            try:
                                data[field] = datetime.strptime(val.strip(), "%d.%m.%Y %H:%M:%S")
                            except Exception:
                                try:
                                    data[field] = datetime.strptime(val.strip(), "%d.%m.%Y")
                                except Exception:
                                    data[field] = None
                        else:
                            data[field] = None
                    else:
                        data[field] = str(val).strip() if val else ""

                # обязательные поля
                required_fields = ['settlement', 'branch', 'support_number']
                missing_required = [f for f in required_fields if not data.get(f)]
                if missing_required:
                    errors.append(f"Строка {row_num}: отсутствуют обязательные поля: {', '.join(missing_required)}")
                    processed_rows += 1
                    continue

                # upsert
                _, is_created = Support.objects.update_or_create(
                    settlement=data.get("settlement"),
                    branch=data.get("branch"),
                    support_number=data.get("support_number"),
                    defaults=data,
                )
                created += 1 if is_created else 0
                updated += 0 if is_created else 1

            except Exception as e:
                errors.append(f"Строка {row_num}: ошибка обработки - {str(e)}")

            processed_rows += 1

        return {
            'success': True,
            'created': created,
            'updated': updated,
            'errors': errors[:10] if errors else [],
            'total_rows': total_rows,
            'processed_rows': processed_rows
        }

    except Exception as e:
        return {
            'success': False,
            'error': f'Ошибка чтения файла: {str(e)}. Убедитесь, что файл не поврежден и имеет правильный формат.'
        }



@login_required
def supports_export_view(request):
    raw = (request.GET.get('status') or '').strip()

    # словари кодов/меток
    allowed_codes = {v for v, _ in FieldSupport.STATUS_CHOICES}
    label_to_code = {label: code for code, label in FieldSupport.STATUS_CHOICES}

    # 'all' или пусто -> показать все (но мы хотим дефолт=processing)
    if raw == 'all':
        status_code = ''
    elif raw in allowed_codes:
        status_code = raw
    elif raw in label_to_code:
        status_code = label_to_code[raw]
    else:
        # ДЕФОЛТ: показываем "В обработке"
        status_code = 'processing'

    qs = FieldSupport.objects.select_related('created_by').order_by('-created_at')
    if status_code:
        qs = qs.filter(status=status_code)

    # Жёстко заданный порядок вкладок
    ordered_status = [
        ('processing', dict(FieldSupport.STATUS_CHOICES)['processing']),
        ('accepted',   dict(FieldSupport.STATUS_CHOICES)['accepted']),
        ('rejected',   dict(FieldSupport.STATUS_CHOICES)['rejected']),
    ]

    context = {
        "supports": qs,
        "status_choices_ordered": ordered_status,
        "current_status": status_code,   # всегда код
        "stats": {
            "total": FieldSupport.objects.count(),
            "processing": FieldSupport.objects.filter(status="processing").count(),
            "accepted": FieldSupport.objects.filter(status="accepted").count(),
            "rejected": FieldSupport.objects.filter(status="rejected").count(),
        },
    }
    return render(request, "main/supports_export.html", context)


@login_required
@require_POST
def update_support_status(request, pk: int):
    """
    Принимает только:
      - status=accepted  (клик по карточке)
      - status=rejected + comment (кнопка Отклонить)
    """
    obj = get_object_or_404(FieldSupport, pk=pk)

    new_status = (request.POST.get("status") or "").strip()
    if new_status not in ("accepted", "rejected"):
        return JsonResponse({"success": False, "error": "Недопустимый статус."}, status=400)

    if new_status == "rejected":
        comment = (request.POST.get("comment") or "").strip()
        if comment:
            obj.comment = comment
        obj.status = "rejected"
        obj.save(update_fields=["status", "comment"])
        return JsonResponse({
            "success": True,
            "new_status": "rejected",
            "new_status_display": obj.get_status_display(),
            "message": f"Опора #{obj.pk} отклонена.",
        })

    # accepted
    if obj.status != "accepted":
        obj.status = "accepted"
        obj.save(update_fields=["status"])
    return JsonResponse({
        "success": True,
        "new_status": "accepted",
        "new_status_display": obj.get_status_display(),
        "message": f"Опора #{obj.pk} принята.",
    })


from django.contrib.auth.decorators import login_required
from django.db import transaction
from django.http import HttpResponse, HttpResponseBadRequest, JsonResponse
from django.shortcuts import get_object_or_404
from io import BytesIO
from zipfile import ZipFile, ZIP_DEFLATED
from django.core.files.storage import default_storage
import os

@login_required
def export_support_photos(request):
    """
    ZIP экспорт фото.

    - POST ids[]=... :
        * берём ТОЛЬКО те ids, что в статусе 'processing'
        * переводим их в 'accepted'
        * пакуем только реально существующие фото
    - GET ?scope=all_processing :
        * берём все объекты в статусе 'processing'
        * пакуем только реально существующие фото
        * после успешной упаковки принудительно переводим упакованные в 'accepted'
    """
    qs = FieldSupport.objects.all()

    mem = BytesIO()
    written = 0
    will_accept_ids = []  # сюда копим только те, у кого реально записали фото в ZIP

    if request.method == "POST":
        ids = request.POST.getlist("ids[]") or request.POST.getlist("ids")
        if not ids:
            return JsonResponse({"success": False, "error": "Не выбраны опоры для экспорта."}, status=400)
        try:
            ids = [int(x) for x in ids]
        except ValueError:
            return HttpResponseBadRequest("Некорректные ids")

        # берём ТОЛЬКО в обработке
        valid_ids = list(
            FieldSupport.objects.filter(pk__in=ids, status="processing").values_list("pk", flat=True)
        )
        if not valid_ids:
            return JsonResponse(
                {"success": False, "error": "Среди выбранных нет опор в обработке."},
                status=400,
            )

        # упаковываем и параллельно собираем, кого действительно приняли (есть фото-файл)
        with ZipFile(mem, "w", ZIP_DEFLATED) as zf:
            for s in qs.filter(pk__in=valid_ids).select_related("created_by"):
                if not s.photo:
                    continue
                storage_path = s.photo.name
                if not default_storage.exists(storage_path):
                    continue
                with default_storage.open(storage_path, "rb") as f:
                    data = f.read()
                base, ext = os.path.splitext(os.path.basename(storage_path))
                owner_part = (s.owner or "").strip().replace(" ", "_")[:40]
                addr_part  = (s.address or "no_address").strip().replace(" ", "_")[:60]
                arcname = f"{s.id}_{owner_part}_{addr_part}{ext or '.jpg'}"
                zf.writestr(arcname, data)
                written += 1
                will_accept_ids.append(s.id)

        if written == 0:
            return JsonResponse(
                {"success": False, "error": "Нет фото для экспорта среди выбранных опор в обработке."},
                status=400,
            )

        # переводим в accepted только реально упакованные
        FieldSupport.objects.filter(pk__in=will_accept_ids, status="processing").update(status="accepted")

    else:  # GET
        scope = request.GET.get("scope")
        if scope != "all_processing":
            return JsonResponse(
                {"success": False, "error": "Укажите scope=all_processing или отправьте ids[] через POST."},
                status=400,
            )

        processing_qs = qs.filter(status="processing")

        with ZipFile(mem, "w", ZIP_DEFLATED) as zf:
            for s in processing_qs.select_related("created_by"):
                if not s.photo:
                    continue
                storage_path = s.photo.name
                if not default_storage.exists(storage_path):
                    continue
                with default_storage.open(storage_path, "rb") as f:
                    data = f.read()
                base, ext = os.path.splitext(os.path.basename(storage_path))
                owner_part = (s.owner or "").strip().replace(" ", "_")[:40]
                addr_part  = (s.address or "no_address").strip().replace(" ", "_")[:60]
                arcname = f"{s.id}_{owner_part}_{addr_part}{ext or '.jpg'}"
                zf.writestr(arcname, data)
                written += 1
                will_accept_ids.append(s.id)

        if written == 0:
            return JsonResponse(
                {"success": False, "error": "Нет фото для экспорта среди опор в обработке."},
                status=400,
            )

        # после успешной упаковки переводим только реально упакованные 'processing' → 'accepted'
        FieldSupport.objects.filter(pk__in=will_accept_ids, status="processing").update(status="accepted")

    mem.seek(0)
    resp = HttpResponse(mem.getvalue(), content_type="application/zip")
    resp["Content-Disposition"] = 'attachment; filename="supports_photos.zip"'
    return resp